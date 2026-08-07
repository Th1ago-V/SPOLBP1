"""Carrinho, finalização, histórico e exportação de compras."""
from io import BytesIO

from flask import Blueprint, abort, flash, redirect, render_template, request, send_file, session, url_for
from openpyxl import Workbook
from openpyxl.styles import Font

from models import compra as compra_model
from models import produto as produto_model


compra_bp = Blueprint("compra", __name__, url_prefix="/compras")


def _carrinho():
    return session.setdefault("carrinho", [])


@compra_bp.get("")
def listar():
    return render_template("compras.html", compras=compra_model.listar())


@compra_bp.get("/nova")
def nova():
    carrinho = _carrinho()
    total = sum(i["preco_centavos"] * i["quantidade"] for i in carrinho)
    return render_template("nova_compra.html", produtos=produto_model.listar(), carrinho=carrinho, total=total)


@compra_bp.post("/carrinho/adicionar")
def adicionar():
    try:
        produto_id = int(request.form.get("produto_id", ""))
        quantidade = int(request.form.get("quantidade", ""))
    except ValueError:
        produto_id, quantidade = 0, 0
    produto = produto_model.buscar(produto_id)
    if produto is None or quantidade <= 0:
        flash("Selecione um produto e uma quantidade válida.", "erro")
    elif quantidade > produto["estoque"]:
        flash("A quantidade supera o estoque disponível.", "erro")
    else:
        carrinho = _carrinho()
        existente = next((i for i in carrinho if i["produto_id"] == produto_id), None)
        nova_quantidade = quantidade + (existente["quantidade"] if existente else 0)
        if nova_quantidade > produto["estoque"]:
            flash("A quantidade total no carrinho supera o estoque.", "erro")
            return redirect(url_for("compra.nova"))
        if existente:
            existente["quantidade"] = nova_quantidade
        else:
            carrinho.append({"produto_id": produto["id"], "nome": produto["nome"],
                             "preco_centavos": produto["preco_centavos"], "quantidade": quantidade})
        session["carrinho"] = carrinho
        flash("Produto adicionado ao carrinho.", "sucesso")
    return redirect(url_for("compra.nova"))


@compra_bp.post("/carrinho/<int:indice>/remover")
def remover(indice):
    carrinho = _carrinho()
    if indice < 0 or indice >= len(carrinho):
        abort(404)
    carrinho.pop(indice)
    session["carrinho"] = carrinho
    return redirect(url_for("compra.nova"))


@compra_bp.post("/carrinho/limpar")
def limpar():
    session.pop("carrinho", None)
    flash("Carrinho esvaziado.", "sucesso")
    return redirect(url_for("compra.nova"))


@compra_bp.post("/finalizar")
def finalizar():
    carrinho = _carrinho()
    if not carrinho:
        flash("O carrinho está vazio.", "erro")
        return redirect(url_for("compra.nova"))
    try:
        compra_id = compra_model.finalizar(session["usuario_id"], carrinho)
    except ValueError as erro:
        flash(str(erro), "erro")
        return redirect(url_for("compra.nova"))
    session.pop("carrinho", None)
    flash("Compra finalizada com sucesso.", "sucesso")
    return redirect(url_for("compra.detalhes", compra_id=compra_id))


@compra_bp.get("/<int:compra_id>")
def detalhes(compra_id):
    compra = compra_model.buscar(compra_id)
    if compra is None:
        abort(404)
    return render_template("detalhes_compra.html", compra=compra,
                           itens=compra_model.listar_itens(compra_id))


@compra_bp.get("/exportar")
def exportar():
    wb = Workbook()
    ws = wb.active
    ws.title = "Compras"
    ws.append(["Código", "Data e hora", "Responsável", "Total (R$)"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for c in compra_model.listar():
        ws.append([c["id"], c["criada_em"], c["usuario_nome"], c["total_centavos"] / 100])
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 25
    for cell in ws["D"][1:]:
        cell.number_format = 'R$ #,##0.00'
    arquivo = BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)
    return send_file(arquivo, as_attachment=True, download_name="compras.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

