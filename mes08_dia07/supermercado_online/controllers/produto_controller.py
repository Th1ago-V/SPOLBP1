"""CRUD e exportação de produtos."""
from io import BytesIO

from flask import Blueprint, abort, flash, redirect, render_template, request, send_file, url_for
from openpyxl import Workbook
from openpyxl.styles import Font

from models import produto as produto_model
from utils import moeda_para_centavos


produto_bp = Blueprint("produto", __name__, url_prefix="/produtos")


def _dados_formulario():
    nome = request.form.get("nome", "").strip()
    fabricante = request.form.get("fabricante", "").strip()
    unidade = request.form.get("unidade", "").strip()
    estoque = int(request.form.get("estoque", "0"))
    preco = moeda_para_centavos(request.form.get("preco", ""))
    if not nome or not fabricante or not unidade or estoque < 0 or preco < 0:
        raise ValueError("Preencha todos os campos com valores válidos.")
    return nome, fabricante, unidade, preco, estoque


@produto_bp.get("")
def listar():
    busca = request.args.get("busca", "")
    return render_template("produtos.html", produtos=produto_model.listar(busca), busca=busca)


@produto_bp.route("/novo", methods=("GET", "POST"))
def novo():
    if request.method == "POST":
        try:
            produto_model.criar(*_dados_formulario())
            flash("Produto cadastrado.", "sucesso")
            return redirect(url_for("produto.listar"))
        except (ValueError, TypeError):
            flash("Preencha preço e estoque com valores válidos.", "erro")
    return render_template("produto_form.html", produto=None)


@produto_bp.route("/<int:produto_id>/editar", methods=("GET", "POST"))
def editar(produto_id):
    produto = produto_model.buscar(produto_id)
    if produto is None:
        abort(404)
    if request.method == "POST":
        try:
            produto_model.atualizar(produto_id, *_dados_formulario())
            flash("Produto atualizado.", "sucesso")
            return redirect(url_for("produto.listar"))
        except (ValueError, TypeError):
            flash("Preencha preço e estoque com valores válidos.", "erro")
    return render_template("produto_form.html", produto=produto)


@produto_bp.post("/<int:produto_id>/excluir")
def excluir(produto_id):
    if produto_model.buscar(produto_id) is None:
        abort(404)
    produto_model.excluir(produto_id)
    flash("Produto excluído. Compras antigas preservam seu nome e preço.", "sucesso")
    return redirect(url_for("produto.listar"))


@produto_bp.get("/exportar")
def exportar():
    wb = Workbook()
    ws = wb.active
    ws.title = "Produtos"
    ws.append(["Código", "Nome", "Fabricante", "Unidade", "Preço (R$)", "Estoque"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for p in produto_model.listar():
        ws.append([p["id"], p["nome"], p["fabricante"], p["unidade"], p["preco_centavos"] / 100, p["estoque"]])
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 22
    for cell in ws["E"][1:]:
        cell.number_format = 'R$ #,##0.00'
    arquivo = BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)
    return send_file(arquivo, as_attachment=True, download_name="produtos.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

